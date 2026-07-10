#!/usr/bin/env python3
"""
Unified TSP comparison script — auto-discovers all tsp_bb versions in solver/.

Algorithms compared:
  1. Concorde                       (exact, state-of-the-art, reference)
  2. Each tsp_bb version            (exact B&B with built-in strategy)

Every tsp_bb version found under solver/ is automatically included in the
comparison.  Just drop a new tsp_bb_YY_MM_DD directory into solver/ and
re-run — no source changes needed.

Each version uses its built-in branching method (no --branch-strategy arg).
Versions that still require --branch-strategy are auto-detected and run with
--branch-strategy smart.

Output: docs/strategy-comparison-results.md  +  .xlsx

Usage:
  python3 tools/compare_strategies.py

To change which instances are tested, edit the INSTANCE_SOURCES block below.
"""

from __future__ import annotations

import subprocess
import sys
import os
import re
import time
import math
import json
import shutil
import tempfile
from pathlib import Path

from typing import cast

try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    from openpyxl.worksheet.worksheet import Worksheet  # type: ignore
    from openpyxl.cell.cell import Cell  # type: ignore
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
CONCORDE = os.path.join(PROJECT_ROOT, "concorde", "TSP", "concorde")
OUT_PATH = os.path.join(PROJECT_ROOT, "docs", "strategy-comparison-results.md")
XLSX_PATH = os.path.join(PROJECT_ROOT, "docs", "strategy-comparison-results.xlsx")
CACHE_PATH = os.path.join(PROJECT_ROOT, "docs", "comparison-cache.json")
CONCORDE_OUT_DIR = os.path.join(PROJECT_ROOT, "TS")

# ── Global settings ──────────────────────────────────────────────────

TIMEOUT = 1800  # seconds per instance per algorithm
DEBUG_INTERVAL = 5000000  # how often tsp_bb prints debug progress
CONCORDE_SEED = 123  # fixed seed for reproducibility

# Fallback strategy for older solver versions that still require --branch-strategy.
LEGACY_STRATEGY = "smart"

# ============================================================
# Instance Sources — edit this block to control which instances
# are included in the comparison. Each entry is a dict:
#
#   path      : directory relative to PROJECT_ROOT (required)
#   recursive : walk subdirectories? True/False (default True)
#   max_dim   : only include instances with n <= max_dim
#               set to None for no size limit
#   skip_dirs : subdirectories to skip when recursive
#               (default: ["_archives", "tsplib"])
#   extensions: file extensions to scan (default: [".txt", ".tsp"])
#
# Examples:
#   {"path": "data/classic/tsplib",  "max_dim": 29}    # TSPLIB, n<30
#   {"path": "data/classic/national","max_dim": 29}    # National, n<30
#   {"path": "data/classic",         "max_dim": 50}    # All classic, n<=50
#   {"path": "examples",             "max_dim": None}   # All examples
#   {"path": "data/classic/vlsi",    "max_dim": 500}   # Small VLSI only
# ============================================================
INSTANCE_SOURCES = [
    # Small project examples (hand-crafted)
    {"path": "examples", "recursive": True, "max_dim": 49,
     "skip_dirs": ["tsplib"]},

    # Classic TSPLIB benchmark — instances with fewer than 30 vertices
    {"path": "data/classic/tsplib", "recursive": False, "max_dim": 59,
     "skip_files": ["ulysses22.tsp"]},
    # result: burma14(14) ulysses16(16) gr17(17) gr21(21) ulysses22(22)
    #         gr24(24) fri26(26) bayg29(29) bays29(29)

    # Classic National benchmark — instances with fewer than 30 vertices
    {"path": "data/classic/national", "recursive": False, "max_dim": 59},
    # result: wi29(29)
]
# ============================================================

INF = float("inf")

DEFAULT_SKIP_DIRS = {"_archives", "tsplib"}
DEFAULT_SKIP_FILES: set[str] = set()
DEFAULT_EXTENSIONS = (".txt", ".tsp")


SOLVER_DIR = os.path.join(PROJECT_ROOT, "solver")

# ── Solver auto-discovery ──────────────────────────────────────────────

def discover_solvers() -> list[tuple[str, str]]:
    """Find all tsp_bb versions in solver/ directory.

    Each subdirectory (e.g. ``tsp_bb_26_07_06``) contains a ``tsp_bb`` binary.
    Returns list of (label, path) tuples sorted by label (chronological).
    Label is the directory name (e.g. ``tsp_bb_26_07_06``).
    """
    if not os.path.isdir(SOLVER_DIR):
        return []
    solvers: list[tuple[str, str]] = []
    for dname in sorted(os.listdir(SOLVER_DIR)):
        dpath = os.path.join(SOLVER_DIR, dname)
        if not os.path.isdir(dpath):
            continue
        if not dname.startswith("tsp_"):
            continue
        binary = os.path.join(dpath, "tsp_bb")
        if os.path.isfile(binary) and os.access(binary, os.X_OK):
            solvers.append((dname, binary))
    return solvers


def needs_strategy_arg(solver_path: str) -> bool:
    """Check if a solver binary still requires --branch-strategy argument.

    Returns True if the --help output mentions --branch-strategy (old versions),
    False otherwise (new versions where BP is built-in).
    """
    try:
        result = subprocess.run(
            [solver_path, "--help"],
            capture_output=True, text=True, timeout=5
        )
        output = result.stderr + result.stdout
        return "--branch-strategy" in output
    except Exception:
        return True  # assume old version if probe fails



# ── Instance Discovery ────────────────────────────────────────────────

def _parse_dimension(filepath: str) -> int | None:
    """Parse DIMENSION from TSPLIB header or first-line-number format."""
    with open(filepath, "r") as f:
        first_line = f.readline().strip()
        tokens = first_line.split()
        if tokens and tokens[0].isdigit():
            n = int(tokens[0])
            if n >= 3:
                return n

    with open(filepath, "r") as f:
        for line in f:
            upper = line.strip().upper()
            if upper.startswith("DIMENSION"):
                line = line.replace(" ", "")  # handle "DIMENSION : 29"
                parts = line.split(":")
                if len(parts) == 2:
                    return int(parts[1].strip())
                parts = line.split()
                if len(parts) >= 2:
                    return int(parts[1])
    return None


def find_instances() -> list[tuple[str, int]]:
    """Discover instances from all configured INSTANCE_SOURCES.

    Returns a list of (absolute_path, dimension) tuples sorted by dimension
    then name.
    """
    instances = []
    seen = set()  # deduplicate by basename

    for cfg in INSTANCE_SOURCES:
        base_path = os.path.join(PROJECT_ROOT, cfg["path"])
        if not os.path.isdir(base_path):
            print(f"Warning: source not found, skipping: {base_path}", file=sys.stderr)
            continue

        recursive = cfg.get("recursive", True)
        max_dim = cfg.get("max_dim", None)
        skip_dirs = set(cfg.get("skip_dirs", DEFAULT_SKIP_DIRS))
        skip_files = set(cfg.get("skip_files", DEFAULT_SKIP_FILES))
        extensions = tuple(cfg.get("extensions", DEFAULT_EXTENSIONS))

        if recursive:
            walker = os.walk(base_path)
        else:
            walker = [(base_path, [], sorted(os.listdir(base_path)))]

        for dirpath, dirnames, filenames in walker:
            if recursive:
                dirnames[:] = [
                    d for d in dirnames
                    if not d.startswith(".") and d not in skip_dirs
                ]

            for fname in sorted(filenames):
                if fname.startswith("."):
                    continue
                if fname.endswith(".gz"):
                    continue
                if fname == "batch.txt":
                    continue
                if not fname.endswith(extensions):
                    continue
                if fname in skip_files:
                    continue

                filepath = os.path.join(dirpath, fname)
                dim = _parse_dimension(filepath)

                if dim is None or dim < 3:
                    continue
                if max_dim is not None and dim > max_dim:
                    continue

                # Deduplicate by filename
                if fname in seen:
                    continue
                seen.add(fname)

                instances.append((filepath, dim))

    # Sort by dimension, then name
    instances.sort(key=lambda x: (x[1], os.path.basename(x[0])))
    return instances


# ── Distance Matrix Loading ───────────────────────────────────────────

def load_distance_matrix(filepath: str) -> tuple[list[list[float]], int]:
    """Load distance matrix from file (raw matrix, TSPLIB coord, or TSPLIB explicit)."""
    with open(filepath, "r") as f:
        content = f.read()

    if "NODE_COORD_SECTION" in content.upper():
        return _load_tsplib_coord(filepath)
    if "EDGE_WEIGHT_SECTION" in content.upper():
        return _load_tsplib_explicit(filepath)

    tokens = content.split()
    n = int(tokens[0])
    matrix = [[0.0] * n for _ in range(n)]
    idx = 1
    for i in range(n):
        for j in range(n):
            matrix[i][j] = _parse_weight(tokens[idx])
            idx += 1
    return matrix, n


def _parse_weight(token: str) -> float:
    lower = token.lower()
    if lower in ("x", "-", "inf", "infinity"):
        return INF
    return float(token)


# Section markers that may appear after the primary data section in TSPLIB files.
_TSBLIB_SECTION_MARKERS = frozenset({
    "NODE_COORD_SECTION", "EDGE_WEIGHT_SECTION",
    "DISPLAY_DATA_SECTION", "TOUR_SECTION",
})


def _load_tsplib_coord(filepath: str) -> tuple[list[list[float]], int]:
    coords = []
    edge_weight_type = "EUC_2D"
    dimension = 0
    section = None

    with open(filepath, "r") as f:
        for line in f:
            line = line.strip()
            upper = line.upper()
            if upper == "NODE_COORD_SECTION":
                section = "coord"
                continue
            if upper == "EOF":
                break
            if section is None:
                if upper.startswith("DIMENSION"):
                    line = line.replace(" ", "")
                    parts = line.split(":")
                    dimension = int(parts[1].strip()) if len(parts) == 2 else int(line.split()[1])
                elif upper.startswith("EDGE_WEIGHT_TYPE"):
                    line = line.replace(" ", "")
                    parts = line.split(":")
                    edge_weight_type = parts[1].strip().upper() if len(parts) == 2 else line.split()[1].upper()
            elif section == "coord":
                if upper in _TSBLIB_SECTION_MARKERS:
                    break  # stop reading coords when another section starts
                parts = line.split()
                if len(parts) >= 3:
                    coords.append((float(parts[1]), float(parts[2])))

    n = dimension if dimension > 0 else len(coords)
    if edge_weight_type == "GEO":
        return _geo_matrix(coords, n), n
    return _euc2d_matrix(coords, n), n


def _euc2d_matrix(coords: list, n: int) -> list[list[float]]:
    matrix = [[0.0] * n for _ in range(n)]
    for i in range(n):
        for j in range(n):
            if i == j:
                continue
            dx = coords[i][0] - coords[j][0]
            dy = coords[i][1] - coords[j][1]
            matrix[i][j] = round(math.sqrt(dx * dx + dy * dy))
    return matrix


def _geo_matrix(coords: list, n: int) -> list[list[float]]:
    PI = math.pi
    matrix = [[0.0] * n for _ in range(n)]

    def to_rad(deg_val):
        d = int(deg_val)
        m = deg_val - d
        return PI * (d + 5.0 * m / 3.0) / 180.0

    for i in range(n):
        for j in range(n):
            if i == j:
                continue
            lat_a = to_rad(coords[i][0])
            lon_a = to_rad(coords[i][1])
            lat_b = to_rad(coords[j][0])
            lon_b = to_rad(coords[j][1])
            q1 = math.cos(lon_a - lon_b)
            q2 = math.cos(lat_a - lat_b)
            q3 = math.cos(lat_a + lat_b)
            arg = 0.5 * ((1.0 + q1) * q2 - (1.0 - q1) * q3)
            arg = max(-1.0, min(1.0, arg))
            matrix[i][j] = int(6378.388 * math.acos(arg) + 1.0)
    return matrix


def _load_tsplib_explicit(filepath: str) -> tuple[list[list[float]], int]:
    dimension = 0
    edge_weight_format = "FULL_MATRIX"
    section = None
    values = []

    with open(filepath, "r") as f:
        for line in f:
            line = line.strip()
            upper = line.upper()
            if upper == "EDGE_WEIGHT_SECTION":
                section = "weights"
                continue
            if upper == "EOF":
                break
            if section is None:
                if upper.startswith("DIMENSION"):
                    line = line.replace(" ", "")
                    parts = line.split(":")
                    dimension = int(parts[1].strip()) if len(parts) == 2 else int(line.split()[1])
                elif upper.startswith("EDGE_WEIGHT_FORMAT"):
                    line = line.replace(" ", "")
                    parts = line.split(":")
                    edge_weight_format = parts[1].strip().upper() if len(parts) == 2 else line.split()[1].upper()
            elif section == "weights":
                if upper in _TSBLIB_SECTION_MARKERS:
                    break  # stop reading weights when another section starts
                for token in line.split():
                    values.append(_parse_weight(token))

    matrix = [[0.0] * dimension for _ in range(dimension)]
    fmt = edge_weight_format
    idx = 0

    if fmt == "FULL_MATRIX":
        for i in range(dimension):
            for j in range(dimension):
                matrix[i][j] = values[idx]
                idx += 1
    elif fmt == "UPPER_ROW":
        for i in range(dimension):
            for j in range(i + 1, dimension):
                matrix[i][j] = matrix[j][i] = values[idx]
                idx += 1
    elif fmt == "LOWER_ROW":
        for i in range(dimension):
            for j in range(i):
                matrix[i][j] = matrix[j][i] = values[idx]
                idx += 1
    elif fmt == "UPPER_DIAG_ROW":
        for i in range(dimension):
            for j in range(i, dimension):
                matrix[i][j] = matrix[j][i] = values[idx]
                idx += 1
    elif fmt == "LOWER_DIAG_ROW":
        for i in range(dimension):
            for j in range(i + 1):
                matrix[i][j] = matrix[j][i] = values[idx]
                idx += 1
    else:
        for i in range(dimension):
            for j in range(dimension):
                if idx < len(values):
                    matrix[i][j] = values[idx]
                    idx += 1
    return matrix, dimension


# ── Format Detection ─────────────────────────────────────────────────

def is_tsplib_format(filepath: str) -> bool:
    """Check if a file is in TSPLIB format (has section markers or header keywords)."""
    with open(filepath, "r") as f:
        content = f.read(4096)
    upper = content.upper()
    for marker in ("EDGE_WEIGHT_SECTION", "NODE_COORD_SECTION",
                    "EDGE_WEIGHT_TYPE", "TYPE:", "NAME:"):
        if marker in upper:
            return True
    return False


# ── Concorde Wrapper ─────────────────────────────────────────────────

def matrix_to_tsplib(filepath: str, out_path: str):
    """Convert a raw matrix-format TSP instance to TSPLIB explicit format.

    Reads the matrix via load_distance_matrix and writes it as TSPLIB
    EDGE_WEIGHT_SECTION / FULL_MATRIX.
    """
    matrix, n = load_distance_matrix(filepath)
    with open(out_path, "w") as f:
        f.write(f"NAME: {os.path.splitext(os.path.basename(filepath))[0]}\n")
        f.write("TYPE: TSP\n")
        f.write(f"DIMENSION: {n}\n")
        f.write("EDGE_WEIGHT_TYPE: EXPLICIT\n")
        f.write("EDGE_WEIGHT_FORMAT: FULL_MATRIX\n")
        f.write("EDGE_WEIGHT_SECTION\n")
        for i in range(n):
            row_vals = []
            for j in range(n):
                val = matrix[i][j]
                if val >= INF / 2:
                    row_vals.append("999999")
                else:
                    row_vals.append(str(int(val)) if val == int(val) else f"{val:.0f}")
            f.write(" ".join(row_vals) + "\n")
        f.write("EOF\n")


def ensure_tsplib_for_concorde(instance_path: str, tmpdir: str) -> str:
    """Ensure instance is in TSPLIB format readable by Concorde.

    If the file is already TSPLIB, returns it unchanged.
    Otherwise converts it to TSPLIB explicit format using Python-native conversion.
    """
    if is_tsplib_format(instance_path):
        return instance_path

    basename = os.path.splitext(os.path.basename(instance_path))[0]
    out_path = os.path.join(tmpdir, basename + ".tsp")
    matrix_to_tsplib(instance_path, out_path)
    return out_path


def parse_concorde_output(stdout: str) -> dict:
    """Extract Concorde solving statistics."""
    stats = {
        "cost": None, "feasible": False, "time": None,
        "bbnodes": None, "status": "unknown",
    }
    for line in stdout.splitlines():
        if "Optimal Solution:" in line:
            m = re.search(r"Optimal Solution:\s*([\d.]+)", line)
            if m:
                stats["cost"] = float(m.group(1))
                stats["feasible"] = True
                stats["status"] = "optimal"
        elif "Total Running Time:" in line:
            m = re.search(r"Total Running Time:\s*([\d.]+)", line)
            if m:
                stats["time"] = float(m.group(1))
        elif "Number of bbnodes:" in line:
            m = re.search(r"Number of bbnodes:\s*(\d+)", line)
            if m:
                stats["bbnodes"] = int(m.group(1))
        elif "ERROR:" in line:
            stats["status"] = "error"
            stats["error_msg"] = line.strip()
    return stats


def parse_concorde_sol(sol_path: str) -> list[int] | None:
    """Parse a Concorde .sol file and return the optimal tour as a list of ints.

    .sol format: first line is the node count, subsequent lines give the
    node indices (0-based) in tour order, space-separated (possibly
    multiple per line). Returns None on failure.
    """
    try:
        with open(sol_path, "r") as f:
            text = f.read()
        tokens = text.split()
        if len(tokens) < 2:
            return None
        n = int(tokens[0])
        tour = [int(v) for v in tokens[1:]]
        if len(tour) == n:
            return tour
        return None
    except (OSError, ValueError):
        return None


def run_concorde(instance_path: str, seed: int = 123) -> dict:
    """Run Concorde on an instance. Returns stats dict with elapsed time and tour.

    Concorde output files (.sol, .sav, etc.) are written to a dedicated
    TS/ subdirectory, never to the project root.
    """
    os.makedirs(CONCORDE_OUT_DIR, exist_ok=True)
    tmpdir = tempfile.mkdtemp(prefix="concorde_", dir=CONCORDE_OUT_DIR)
    try:
        tsplib_path = ensure_tsplib_for_concorde(instance_path, tmpdir)

        # Copy input into tmpdir so Concorde writes output files inside it.
        if Path(tsplib_path).parent != Path(tmpdir):
            dest = Path(tmpdir) / Path(tsplib_path).name
            shutil.copy2(tsplib_path, dest)
            tsplib_path = str(dest)

        # Explicitly set the output file so Concorde writes the tour into
        # tmpdir even when it would otherwise default to the project root.
        out_name = Path(tsplib_path).stem + ".sol"
        out_file = os.path.join(tmpdir, out_name)

        cmd = [
            CONCORDE,
            "-s", str(seed),
            "-o", out_file,
            tsplib_path,
        ]
        t0 = time.perf_counter()
        try:
            # Run with cwd=tmpdir so any extra Concorde working files
            # (.sav, .pul, .mas) land there as well.
            result = subprocess.run(cmd, capture_output=True, text=True,
                                    timeout=TIMEOUT, cwd=tmpdir)
            elapsed = time.perf_counter() - t0
            stats = parse_concorde_output(result.stdout)
            stats["elapsed"] = elapsed
            stats["timeout"] = False
            stats["error"] = None
            stats["stderr"] = result.stderr
            stats["tour"] = None
            stats["tour_list"] = None

            # Locate the .sol file Concorde wrote inside tmpdir
            sol_glob = list(Path(tmpdir).glob("*.sol"))
            if sol_glob:
                tour = parse_concorde_sol(str(sol_glob[0]))
                if tour:
                    stats["tour"] = " -> ".join(str(v) for v in tour)
                    stats["tour_list"] = tour
        except subprocess.TimeoutExpired:
            stats = _empty_stats_concorde()
            stats["elapsed"] = TIMEOUT
            stats["timeout"] = True
        except Exception as e:
            stats = _empty_stats_concorde()
            stats["elapsed"] = time.perf_counter() - t0
            stats["error"] = str(e)
        return stats
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def _empty_stats_concorde() -> dict:
    return {
        "cost": None, "feasible": False, "time": None, "bbnodes": None,
        "status": "unknown", "elapsed": 0.0, "timeout": False,
        "error": None, "stderr": "",
        "tour": None, "tour_list": None,
    }


# ── tsp_bb Solver Wrapper ────────────────────────────────────────────

def parse_tspbb_stats(stdout: str) -> dict:
    """Extract solver statistics from tsp_bb human-readable output."""
    stats = {
        "cost": None, "feasible": False, "root_lb": None, "init_ub": None,
        "nodes_created": None, "nodes_expanded": None,
        "pruned_bound": None, "pruned_infeasible": None,
        "tour": None, "tour_list": None,
    }
    for line in stdout.splitlines():
        if "Optimal cost:" in line:
            m = re.search(r"Optimal cost:\s*([\d.]+)", line)
            if m:
                stats["cost"] = float(m.group(1))
                stats["feasible"] = True
        elif "Root lower bound:" in line:
            m = re.search(r"Root lower bound:\s*([\d.]+)", line)
            if m:
                stats["root_lb"] = float(m.group(1))
        elif "Initial upper bound:" in line:
            m = re.search(r"Initial upper bound:\s*([\d.]+)", line)
            if m:
                stats["init_ub"] = float(m.group(1))
        elif "Nodes created:" in line:
            m = re.search(r"Nodes created:\s*(\d+)", line)
            if m:
                stats["nodes_created"] = int(m.group(1))
        elif "Nodes expanded:" in line:
            m = re.search(r"Nodes expanded:\s*(\d+)", line)
            if m:
                stats["nodes_expanded"] = int(m.group(1))
        elif "Pruned by bound:" in line:
            m = re.search(r"Pruned by bound:\s*(\d+)", line)
            if m:
                stats["pruned_bound"] = int(m.group(1))
        elif "Pruned infeasible:" in line:
            m = re.search(r"Pruned infeasible:\s*(\d+)", line)
            if m:
                stats["pruned_infeasible"] = int(m.group(1))
        elif "Tour:" in line:
            m = re.search(r"Tour:\s*(.*)", line)
            if m:
                tour_str = m.group(1).strip()
                stats["tour"] = tour_str
                parts = tour_str.replace("->", " ").split()
                try:
                    stats["tour_list"] = [int(x) for x in parts if x.lstrip("-").isdigit()]
                except ValueError:
                    pass
    return stats


def run_tspbb(instance_path: str, binary: str, add_strategy_arg: bool = False) -> dict:
    """Run tsp_bb solver. Returns stats dict with elapsed time."""
    cmd = [binary]
    if add_strategy_arg:
        cmd += ["--branch-strategy", LEGACY_STRATEGY]
    cmd += [
        "--exact-max-n", "100",
        "--debug",
        "--debug-interval", str(DEBUG_INTERVAL),
        instance_path,
    ]
    t0 = time.perf_counter()
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=TIMEOUT)
        elapsed = time.perf_counter() - t0
        stats = parse_tspbb_stats(result.stdout)
        stats["elapsed"] = elapsed
        stats["timeout"] = False
        stats["error"] = None
        stats["stderr"] = result.stderr
    except subprocess.TimeoutExpired:
        stats = _empty_stats_tspbb()
        stats["elapsed"] = TIMEOUT
        stats["timeout"] = True
    except Exception as e:
        stats = _empty_stats_tspbb()
        stats["elapsed"] = time.perf_counter() - t0
        stats["error"] = str(e)
    return stats


def _empty_stats_tspbb() -> dict:
    return {
        "cost": None, "feasible": False, "root_lb": None, "init_ub": None,
        "nodes_created": None, "nodes_expanded": None,
        "pruned_bound": None, "pruned_infeasible": None,
        "tour": None, "tour_list": None,
        "elapsed": 0.0, "timeout": False, "error": None, "stderr": "",
    }


# ── Persistent Cache ─────────────────────────────────────────────────
#
# Cache file (JSON) stores:
#   {
#     "concorde": { "<rel_path>": { cost, elapsed, bbnodes, tour, tour_list } },
#     "solver_results": {
#       "<solver_label>": { "<rel_path>": { cost, elapsed, nodes_created, ... } }
#     },
#     "solver_timeouts": {
#       "<solver_label>": { "<rel_path>": true }
#     }
#   }
#
# On startup the cache is loaded so that:
#   - Instances that previously timed out (30 min) for a solver are skipped.
#   - Instances that already have a successful result for a solver are reused.
#   - Concorde results are also reused (they are deterministic with a fixed seed).
#
# The XLSX file is regenerated each run from the cache + newly computed data.

def load_cache() -> dict:
    """Load the comparison cache from disk. Returns empty dict if not found."""
    if not os.path.isfile(CACHE_PATH):
        return {}
    try:
        with open(CACHE_PATH, "r") as f:
            cache = json.load(f)
        if not isinstance(cache, dict):
            return {}
        # Ensure expected sub-dicts exist
        cache.setdefault("concorde", {})
        cache.setdefault("solver_results", {})
        cache.setdefault("solver_timeouts", {})
        return cache
    except (json.JSONDecodeError, OSError) as e:
        print(f"Warning: failed to load cache ({e}), starting fresh.", file=sys.stderr)
        return {"concorde": {}, "solver_results": {}, "solver_timeouts": {}}


def save_cache(cache: dict) -> None:
    """Persist the cache to disk atomically."""
    os.makedirs(os.path.dirname(CACHE_PATH), exist_ok=True)
    tmp = CACHE_PATH + ".tmp"
    try:
        with open(tmp, "w") as f:
            json.dump(cache, f, indent=2, sort_keys=True)
        os.replace(tmp, CACHE_PATH)
    except OSError as e:
        print(f"Warning: failed to save cache: {e}", file=sys.stderr)


def get_cached_concorde(cache: dict, rel_path: str) -> dict | None:
    """Return cached Concorde stats for an instance, or None."""
    return cache.get("concorde", {}).get(rel_path)


def store_concorde_result(cache: dict, rel_path: str, stats: dict) -> None:
    """Store Concorde stats in the cache."""
    entry = {
        "cost": stats.get("cost"),
        "feasible": stats.get("feasible", False),
        "elapsed": stats.get("elapsed", 0),
        "bbnodes": stats.get("bbnodes"),
        "tour": stats.get("tour"),
        "tour_list": stats.get("tour_list"),
        "error": stats.get("error"),
        "timeout": stats.get("timeout", False),
    }
    # Strip non-serialisable fields and huge strings
    entry.pop("stderr", None)
    cache.setdefault("concorde", {})[rel_path] = entry


def is_solver_timeout_cached(cache: dict, solver_label: str, rel_path: str) -> bool:
    """Check whether a solver+instance combination already timed out."""
    return rel_path in cache.get("solver_timeouts", {}).get(solver_label, {})


def get_cached_solver_result(cache: dict, solver_label: str, rel_path: str) -> dict | None:
    """Return cached solver stats, or None."""
    return cache.get("solver_results", {}).get(solver_label, {}).get(rel_path)


def store_solver_timeout(cache: dict, solver_label: str, rel_path: str) -> None:
    """Record a solver timeout in the cache."""
    cache.setdefault("solver_timeouts", {}).setdefault(solver_label, {})[rel_path] = True


def store_solver_result(cache: dict, solver_label: str, rel_path: str, stats: dict) -> None:
    """Store a successful solver result in the cache."""
    entry = {
        "cost": stats.get("cost"),
        "feasible": stats.get("feasible", False),
        "elapsed": stats.get("elapsed", 0),
        "root_lb": stats.get("root_lb"),
        "init_ub": stats.get("init_ub"),
        "nodes_created": stats.get("nodes_created"),
        "nodes_expanded": stats.get("nodes_expanded"),
        "pruned_bound": stats.get("pruned_bound"),
        "pruned_infeasible": stats.get("pruned_infeasible"),
        "tour": stats.get("tour"),
        "tour_list": stats.get("tour_list"),
        "timeout": False,
    }
    cache.setdefault("solver_results", {}).setdefault(solver_label, {})[rel_path] = entry

    # If we previously recorded a timeout for this combination, clear it
    # (a successful run supersedes the timeout).
    to_dict = cache.get("solver_timeouts", {}).get(solver_label, {})
    if rel_path in to_dict:
        del to_dict[rel_path]


def cached_stats_to_result(cached: dict) -> dict:
    """Convert a cached solver entry back to the stats dict shape expected
    by downstream code."""
    return {
        "cost": cached.get("cost"),
        "feasible": cached.get("feasible", False),
        "root_lb": cached.get("root_lb"),
        "init_ub": cached.get("init_ub"),
        "nodes_created": cached.get("nodes_created"),
        "nodes_expanded": cached.get("nodes_expanded"),
        "pruned_bound": cached.get("pruned_bound"),
        "pruned_infeasible": cached.get("pruned_infeasible"),
        "tour": cached.get("tour"),
        "tour_list": cached.get("tour_list"),
        "elapsed": cached.get("elapsed", 0),
        "timeout": cached.get("timeout", False),
        "error": None,
        "stderr": "",
    }


def cached_concorde_to_stats(cached: dict) -> dict:
    """Convert a cached Concorde entry back to the stats dict shape."""
    return {
        "cost": cached.get("cost"),
        "feasible": cached.get("feasible", False),
        "time": None,
        "bbnodes": cached.get("bbnodes"),
        "status": "optimal" if cached.get("feasible") else "unknown",
        "elapsed": cached.get("elapsed", 0),
        "timeout": cached.get("timeout", False),
        "error": cached.get("error"),
        "stderr": "",
        "tour": cached.get("tour"),
        "tour_list": cached.get("tour_list"),
    }


# ── Formatting ────────────────────────────────────────────────────────

def fmt_cost(val) -> str:
    if val is None or val >= INF / 2:
        return "-"
    return f"{val:.0f}"


def fmt_time(sec: float | None) -> str:
    if sec is None:
        return "-"
    if sec < 0.001:
        return "<1ms"
    if sec < 1:
        return f"{sec * 1000:.0f}ms"
    if sec < 60:
        return f"{sec:.1f}s"
    if sec >= 3600:
        return f"{sec / 3600:.1f}h"
    return f"{sec / 60:.1f}m"


def fmt_val(val, template="{}") -> str:
    if val is None:
        return "-"
    return template.format(val)


def cost_match(alg_cost: float | None, ref_cost: float | None) -> str:
    """Return emoji match indicator comparing against reference cost."""
    if alg_cost is None or ref_cost is None:
        return "-"
    if ref_cost >= INF / 2:
        return "-"
    if abs(alg_cost - ref_cost) < 1e-6:
        return ":white_check_mark:"
    return ":x:"


def _build_source_description() -> str:
    """Build a human-readable description of the configured instance sources."""
    parts = []
    for cfg in INSTANCE_SOURCES:
        path = cfg["path"]
        max_dim = cfg.get("max_dim")
        dim_str = f"n <= {max_dim}" if max_dim is not None else "all sizes"
        parts.append(f"`{path}` ({dim_str})")
    return ", ".join(parts)


# ── Main ──────────────────────────────────────────────────────────────

def main():
    # ── Discover solvers ──
    solvers = discover_solvers()
    if not solvers:
        print("Error: no tsp_* solver binaries found in build/", file=sys.stderr)
        print("Build one first:  cd build && cmake .. && make", file=sys.stderr)
        sys.exit(1)

    print(f"Discovered {len(solvers)} solver version(s):", file=sys.stderr)
    for label, path in solvers:
        needs_arg = needs_strategy_arg(path)
        desc = "legacy (--branch-strategy)" if needs_arg else "BP built-in"
        print(f"  {label}  ({path})  {desc}", file=sys.stderr)
    print(file=sys.stderr)

    # Validate prerequisites
    errors = []
    if not os.path.isfile(CONCORDE):
        errors.append(f"Concorde not found: {CONCORDE}")
    for label, path in solvers:
        if not os.path.isfile(path):
            errors.append(f"Solver not found: {path}")
    if errors:
        for e in errors:
            print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

    # Build flat solver list — each version runs once with its built-in method.
    solver_entries: list[tuple[str, str, bool]] = []  # (label, path, needs_arg)
    for label, path in solvers:
        needs_arg = needs_strategy_arg(path)
        solver_entries.append((label, path, needs_arg))
        desc = "legacy (--branch-strategy smart)" if needs_arg else "BP built-in"
        print(f"  {label}: {desc}", file=sys.stderr)

    # Summary keys: solver label
    def _summary_key(solver_label: str) -> str:
        return solver_label

    all_summary_keys = [label for label, _, _ in solver_entries]
    summary: dict[str, dict] = {
        key: {"ok": 0, "timeout": 0, "error": 0, "wrong_cost": 0,
              "total_time": 0.0, "total_nodes": 0, "count_nodes": 0}
        for key in all_summary_keys
    }

    instances = find_instances()
    if not instances:
        print("Error: no instances found matching the configured sources", file=sys.stderr)
        print("Check INSTANCE_SOURCES in the script header.", file=sys.stderr)
        sys.exit(1)

    source_desc = _build_source_description()
    print(f"Found {len(instances)} instances from {len(INSTANCE_SOURCES)} source(s):", file=sys.stderr)
    for p, d in instances:
        print(f"  n={d:3d}  {os.path.relpath(p, PROJECT_ROOT)}", file=sys.stderr)
    print(file=sys.stderr)

    # ── Load persistent cache ──
    cache = load_cache()
    cached_concorde = sum(1 for v in cache.get("concorde", {}).values() if v.get("feasible"))
    cached_solver = sum(len(v) for v in cache.get("solver_results", {}).values())
    cached_timeouts = sum(len(v) for v in cache.get("solver_timeouts", {}).values())
    if cached_concorde or cached_solver or cached_timeouts:
        print(f"Cache loaded: {cached_concorde} Concorde, {cached_solver} solver results, "
              f"{cached_timeouts} timeout records", file=sys.stderr)
        print(file=sys.stderr)

    # ── Build markdown header ──
    lines: list[str] = []
    lines.append("# TSP Algorithm Comparison")
    lines.append("")
    lines.append("## Algorithms Compared")
    lines.append("")
    lines.append("| Algorithm | Description |")
    lines.append("|---|---|")
    lines.append("| **Concorde** | State-of-the-art Concorde TSP solver (exact, with QSopt LP) — **reference** |")
    for label, path, needs_arg in solver_entries:
        method = "smart branching (legacy)" if needs_arg else "BP (Branch Partitioning)"
        lines.append(f"| **{label}** | Branch & Bound solver — {method} |")
    lines.append("")
    lines.append(f"**Instances:** {len(instances)} from {source_desc}  ")
    lines.append(f"**Timeout:** {TIMEOUT}s ({TIMEOUT // 3600}h) per method per instance  ")
    lines.append(f"**Reference:** Concorde exact solver  ")
    lines.append(f"**Solver versions:** {', '.join(f'`{l}`' for l, _, _ in solver_entries)}  ")
    lines.append("")
    lines.append("---")
    lines.append("")

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)

    def flush():
        with open(OUT_PATH, "w") as f:
            f.write("\n".join(lines))

    flush()

    per_instance: list[dict] = []

    for idx, (path, dim) in enumerate(instances, 1):
        rel_path = os.path.relpath(path, PROJECT_ROOT)
        print(f"\n[{idx}/{len(instances)}] {rel_path} (n={dim})", file=sys.stderr)

        # ── Concorde (reference) ──
        cached_cc = get_cached_concorde(cache, rel_path)
        if cached_cc is not None:
            concorde_stats = cached_concorde_to_stats(cached_cc)
            ref_cost = concorde_stats.get("cost") if concorde_stats.get("feasible") else None
            ref_feasible = concorde_stats.get("feasible", False)
            print(f"  Concorde: (cached) cost={ref_cost} time={fmt_time(concorde_stats.get('elapsed', 0))}", file=sys.stderr)
        else:
            print(f"  Concorde...", file=sys.stderr, end=" ", flush=True)
            concorde_stats = run_concorde(path, CONCORDE_SEED)
            store_concorde_result(cache, rel_path, concorde_stats)
            save_cache(cache)
            ref_cost = concorde_stats.get("cost") if concorde_stats.get("feasible") else None
            ref_feasible = concorde_stats.get("feasible", False)
            if concorde_stats["timeout"]:
                print(f"TIMEOUT", file=sys.stderr)
            elif concorde_stats["error"]:
                print(f"ERROR: {concorde_stats['error']}", file=sys.stderr)
            else:
                c_time = concorde_stats.get("elapsed", 0)
                c_bb = concorde_stats.get("bbnodes")
                if ref_cost is not None:
                    print(f"cost={ref_cost:.0f} time={fmt_time(c_time)} bbnodes={c_bb}", file=sys.stderr)
                else:
                    print(f"no solution time={fmt_time(c_time)}", file=sys.stderr)

        # ── Run all solver versions ──
        all_results: dict[str, dict] = {}  # keyed by solver_label
        for s_label, s_path, needs_arg in solver_entries:
            # Check timeout cache first — skip immediately if known to time out.
            if is_solver_timeout_cached(cache, s_label, rel_path):
                st = _empty_stats_tspbb()
                st["elapsed"] = TIMEOUT
                st["timeout"] = True
                all_results[s_label] = st
                print(f"  {s_label}: (cached TIMEOUT — skipped)", file=sys.stderr)
                continue

            # Check result cache — reuse if we already have a successful run.
            cached_sr = get_cached_solver_result(cache, s_label, rel_path)
            if cached_sr is not None:
                st = cached_stats_to_result(cached_sr)
                all_results[s_label] = st
                print(f"  {s_label}: ", file=sys.stderr, end="")
                _print_solver_status(s_label, st, ref_cost)
                print(f"           (cached)", file=sys.stderr)
                continue

            # Run the solver.
            print(f"  {s_label}: ", file=sys.stderr, end="", flush=True)
            st = run_tspbb(path, binary=s_path, add_strategy_arg=needs_arg)
            all_results[s_label] = st
            _print_solver_status(s_label, st, ref_cost)

            # Persist to cache.
            if st["timeout"]:
                store_solver_timeout(cache, s_label, rel_path)
            elif not st["error"]:
                store_solver_result(cache, s_label, rel_path, st)
            save_cache(cache)

        # ── Update summary ──
        for s_label, s_path, needs_arg in solver_entries:
            skey = _summary_key(s_label)
            st = all_results[s_label]
            if st["timeout"]:
                summary[skey]["timeout"] += 1
            elif st["error"]:
                summary[skey]["error"] += 1
            elif st.get("feasible"):
                if ref_feasible and abs(st["cost"] - ref_cost) > 1e-6:
                    summary[skey]["wrong_cost"] += 1
                else:
                    summary[skey]["ok"] += 1
                summary[skey]["total_time"] += st["elapsed"]
                nodes = st.get("nodes_created")
                if nodes is not None:
                    summary[skey]["total_nodes"] += nodes
                    summary[skey]["count_nodes"] += 1
            else:
                summary[skey]["ok"] += 1
                summary[skey]["total_time"] += st["elapsed"]

        # ── Instance output ──
        lines.append(f"## {idx}. `{rel_path}` (n={dim})")
        lines.append("")

        if ref_feasible:
            lines.append(f"**Concorde optimal cost:** {fmt_cost(ref_cost)}  ")
            lines.append(f"**Concorde time:** {fmt_time(concorde_stats['elapsed'])}  ")
        else:
            lines.append(f"**Concorde:** failed to find optimal solution  ")
        lines.append("")

        # Concorde tour (reference)
        concorde_tour = concorde_stats.get("tour_list")
        if concorde_tour:
            t_str = " -> ".join(str(v) for v in concorde_tour) + f" -> {concorde_tour[0]}"
            lines.append(f"**Concorde (reference) tour:** `{t_str}`  ")
            lines.append("")

        # ── Solver comparison table ──
        lines.append("### Results")
        lines.append("")
        header_cols = ["Solver", "Cost", "Time", "Nodes Created", "Nodes Expanded",
                       "Pruned(Bound)", "Pruned(Infeas)", "Match Ref"]
        lines.append("| " + " | ".join(header_cols) + " |")
        lines.append("|" + "|".join(["---"] * len(header_cols)) + "|")

        for s_label, _, _ in solver_entries:
            st = all_results.get(s_label)
            if st is None:
                continue
            if st["timeout"]:
                lines.append(f"| {s_label} | TIMEOUT | {fmt_time(st['elapsed'])} | - | - | - | - | - |")
            elif st["error"]:
                lines.append(f"| {s_label} | ERROR | - | - | - | - | - | - |")
            else:
                cost_s = fmt_cost(st.get("cost"))
                match = cost_match(st.get("cost"), ref_cost)
                lines.append(
                    f"| {s_label} | {cost_s} | {fmt_time(st.get('elapsed', 0))} "
                    f"| {fmt_val(st.get('nodes_created'))} | {fmt_val(st.get('nodes_expanded'))} "
                    f"| {fmt_val(st.get('pruned_bound'))} | {fmt_val(st.get('pruned_infeasible'))} "
                    f"| {match} |"
                )
        lines.append("")

        # ── Tours ──
        tour_lines = []
        # Concorde tour
        if concorde_tour and len(concorde_tour) > 1:
            t_str = " -> ".join(str(v) for v in concorde_tour) + f" -> {concorde_tour[0]}"
            tour_lines.append(f"- **Concorde (reference):** `{t_str}` cost={fmt_cost(concorde_stats.get('cost'))}")

        for s_label, _, _ in solver_entries:
            st = all_results.get(s_label)
            if st and st.get("tour_list") and len(st["tour_list"]) > 1:
                tour_list = list(dict.fromkeys(st["tour_list"]))
                t_str = " -> ".join(str(v) for v in tour_list) + f" -> {tour_list[0]}"
                cost_info = ""
                if ref_feasible and st.get("feasible"):
                    if abs(st["cost"] - ref_cost) < 1e-6:
                        if concorde_tour and tour_list == concorde_tour:
                            cost_info = " (=ref, same tour)"
                        else:
                            cost_info = " (=ref, different tour)"
                    else:
                        cost_info = " (DIFFERS from ref)"
                tour_lines.append(
                    f"- **{s_label}:** `{t_str}` cost={fmt_cost(st.get('cost'))}{cost_info}"
                )

        if tour_lines:
            lines.append("**Tours found:**")
            lines.extend(tour_lines)
            lines.append("")

        # ── Collect per-instance data for Excel ──
        def _alg_cost_val(st):
            if st["timeout"] or st["error"] or not st.get("feasible"):
                return None
            return st["cost"]

        def _alg_time_val(st):
            if st["timeout"] or st["error"]:
                return None
            return st["elapsed"]

        def _alg_nodes_val(st):
            return st.get("nodes_created")

        pi_entry: dict = {
            "idx": idx,
            "name": rel_path,
            "n": dim,
            "concorde_cost": fmt_cost(concorde_stats.get("cost")) if concorde_stats.get("feasible") else (
                "ERR" if concorde_stats["error"] else "TO" if concorde_stats["timeout"] else "-"),
            "concorde_time": fmt_time(concorde_stats["elapsed"]),
            "_concorde_cost": _alg_cost_val(concorde_stats),
            "_concorde_time": _alg_time_val(concorde_stats),
            "_concorde_nodes": concorde_stats.get("bbnodes"),
        }
        for s_label, _, _ in solver_entries:
            st = all_results[s_label]
            pi_entry[f"{s_label}_cost_str"] = (
                fmt_cost(st.get("cost")) if st.get("feasible") else (
                    "ERR" if st["error"] else "TO" if st["timeout"] else "-"))
            pi_entry[f"{s_label}_time_str"] = fmt_time(st["elapsed"])
            pi_entry[f"_{s_label}_cost"] = _alg_cost_val(st)
            pi_entry[f"_{s_label}_time"] = _alg_time_val(st)
            pi_entry[f"_{s_label}_nodes"] = _alg_nodes_val(st)
        per_instance.append(pi_entry)

        lines.append("---")
        lines.append("")
        flush()

    # ── Summary ──
    lines.append("")
    lines.append("## Summary")
    lines.append("")

    sum_header = ["Solver", "Solved", "Timeout", "Error", "Wrong Cost",
                  "Total Time", "Avg Time", "Avg Nodes", "Total Nodes"]
    lines.append("| " + " | ".join(sum_header) + " |")
    lines.append("|" + "|".join(["---"] * len(sum_header)) + "|")

    for s_label, _, _ in solver_entries:
        skey = _summary_key(s_label)
        s = summary.get(skey, {})
        ok = s.get("ok", 0)
        timeout = s.get("timeout", 0)
        error = s.get("error", 0)
        wrong = s.get("wrong_cost", 0)
        total_time = s.get("total_time", 0.0)
        avg_time = total_time / ok if ok > 0 else 0
        total_nodes = s.get("total_nodes", 0)
        count_nodes = s.get("count_nodes", 0)
        avg_nodes = total_nodes / count_nodes if count_nodes > 0 else 0
        lines.append(
            f"| {s_label} | {ok} | {timeout} | {error} | {wrong} | {fmt_time(total_time)} "
            f"| {fmt_time(avg_time)} | {avg_nodes:.0f} | {total_nodes} |"
        )
    lines.append("")

    # Cross-version speedup (if multiple solvers)
    if len(solvers) >= 2:
        lines.append("### Cross-Version Speedup")
        lines.append("")
        lines.append("Speedup of each solver relative to the first (baseline) solver, "
                      "computed from total solve time over all instances.")
        lines.append("")
        baseline_label = solvers[0][0]
        base_skey = _summary_key(baseline_label)
        base_time = summary[base_skey]["total_time"]
        base_nodes = (summary[base_skey]["total_nodes"] / summary[base_skey]["count_nodes"]
                      if summary[base_skey]["count_nodes"] > 0 else 0)

        sp_header = ["Solver", "Total Time", "Speedup vs Baseline", "Avg Nodes", "Node Ratio"]
        lines.append("| " + " | ".join(sp_header) + " |")
        lines.append("|" + "|".join(["---"] * len(sp_header)) + "|")

        for s_label, _, _ in solver_entries:
            skey = _summary_key(s_label)
            s = summary[skey]
            total_t = s["total_time"]
            avg_n = (s["total_nodes"] / s["count_nodes"]
                     if s["count_nodes"] > 0 else 0)
            if total_t > 0 and base_time > 0:
                sp = base_time / total_t
                sp_str = f"{sp:.2f}x" if sp >= 1 else f"{1/sp:.2f}x slower"
            else:
                sp_str = "-"
            node_ratio_str = ""
            if base_nodes > 0 and avg_n > 0:
                node_ratio_str = f"{avg_n / base_nodes:.2f}x"
            lines.append(
                f"| {s_label} | {fmt_time(total_t)} | {sp_str} "
                f"| {avg_n:.0f} | {node_ratio_str} |"
            )
        lines.append("")

    # Per-instance matrix
    lines.append("### Per-Instance Results Matrix")
    lines.append("")
    matrix_header = ["#", "Instance", "n", "Concorde Cost"]
    for s_label, _, _ in solver_entries:
        matrix_header.append(f"{s_label} Cost")
        matrix_header.append(f"{s_label} Time")
    lines.append("| " + " | ".join(matrix_header) + " |")
    lines.append("|" + "|".join(["---"] * len(matrix_header)) + "|")

    for pi in per_instance:
        row = [str(pi["idx"]), f"`{pi['name']}`", str(pi["n"]), str(pi["concorde_cost"])]
        for s_label, _, _ in solver_entries:
            row.append(str(pi.get(f"{s_label}_cost_str", "-")))
            row.append(str(pi.get(f"{s_label}_time_str", "-")))
        lines.append("| " + " | ".join(row) + " |")
    lines.append("")

    flush()

    # Excel export
    export_xlsx(per_instance, summary, solver_entries, XLSX_PATH)

    print(f"\nResults written to: {OUT_PATH}", file=sys.stderr)


def _print_solver_status(_label: str, stats: dict, ref_cost: float | None):
    """Print a one-line solver status to stderr."""
    if stats["timeout"]:
        print(f"TIMEOUT", file=sys.stderr)
    elif stats["error"]:
        print(f"ERROR: {stats['error']}", file=sys.stderr)
    elif stats["feasible"]:
        cost = stats["cost"]
        expanded = stats.get("nodes_expanded")
        if ref_cost is not None and ref_cost < INF / 2:
            tag = "OK (=ref)" if abs(cost - ref_cost) < 1e-6 else f"WRONG (ref={ref_cost:.0f})"
        else:
            tag = "OK (no reference)"
        print(f"cost={cost:.0f} {tag} time={fmt_time(stats['elapsed'])} expanded={expanded}", file=sys.stderr)
    else:
        print(f"INFEASIBLE", file=sys.stderr)


# ── Excel Export ────────────────────────────────────────────────────────

def export_xlsx(per_instance: list[dict], summary: dict,
                solver_entries: list[tuple[str, str, bool]],
                xlsx_path: str):
    """Write per-instance results and summary to an Excel workbook."""
    if not HAS_OPENPYXL:
        print("Warning: openpyxl not installed. Skipping Excel export. "
              "Install with: pip install openpyxl", file=sys.stderr)
        return

    wb = openpyxl.Workbook()

    # ── Styles ──
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws: Worksheet, row: int, ncols: int) -> None:
        for col in range(1, ncols + 1):
            cell = cast(Cell, ws.cell(row=row, column=col))
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_data_cell(cell: Cell) -> None:
        cell.alignment = cell_align
        cell.border = thin_border

    solver_labels = [label for label, _, _ in solver_entries]

    # ═══════════════════════════════════════════════════════════════
    # Per-Instance Results sheet
    # ═══════════════════════════════════════════════════════════════
    ws = cast(Worksheet, wb.create_sheet("Per-Instance Results"))

    headers = ["#", "Instance", "n", "Concorde Cost", "Concorde Time (s)", "Concorde Nodes"]
    for s_label in solver_labels:
        headers.append(f"{s_label} Cost")
        headers.append(f"{s_label} Time (s)")
        headers.append(f"{s_label} Nodes")

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    for r, pi in enumerate(per_instance, 2):
        values: list = [
            pi["idx"], pi["name"], pi["n"],
            pi.get("_concorde_cost"), pi.get("_concorde_time"), pi.get("_concorde_nodes"),
        ]
        for s_label in solver_labels:
            values.append(pi.get(f"_{s_label}_cost"))
            values.append(pi.get(f"_{s_label}_time"))
            values.append(pi.get(f"_{s_label}_nodes"))
        for c, v in enumerate(values, 1):
            cell = cast(Cell, ws.cell(row=r, column=c))
            if v is None:
                cell.value = "-"
            elif isinstance(v, float):
                cell.value = round(v, 2)
            else:
                cell.value = str(v)
            style_data_cell(cell)

    base_widths = [5, 35, 6, 14, 14, 14]
    solver_widths = [14, 14, 14]
    all_widths = base_widths + solver_widths * len(solver_labels)
    for c, w in enumerate(all_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "D2"

    # ═══════════════════════════════════════════════════════════════
    # Summary sheet
    # ═══════════════════════════════════════════════════════════════
    ws_summary = cast(Worksheet, wb.create_sheet("Summary"))

    sum_headers = [
        "Solver", "Solved", "Timeout", "Error", "Wrong Cost",
        "Total Time (s)", "Avg Time (s)", "Avg Nodes", "Total Nodes",
    ]
    for c, h in enumerate(sum_headers, 1):
        ws_summary.cell(row=1, column=c, value=h)
    style_header(ws_summary, 1, len(sum_headers))

    row = 2
    for s_label in solver_labels:
        s = summary.get(s_label, {})
        ok = s.get("ok", 0)
        total_time = s.get("total_time", 0.0)
        avg_time = total_time / ok if ok > 0 else 0
        total_nodes = s.get("total_nodes", 0)
        count_nodes = s.get("count_nodes", 0)
        avg_nodes = total_nodes / count_nodes if count_nodes > 0 else 0
        values = [
            s_label, ok, s.get("timeout", 0),
            s.get("error", 0), s.get("wrong_cost", 0),
            round(total_time, 2), round(avg_time, 2),
            round(avg_nodes, 0), total_nodes,
        ]
        for c, v in enumerate(values, 1):
            cell = cast(Cell, ws_summary.cell(row=row, column=c))
            cell.value = v
            style_data_cell(cell)
        row += 1

    sum_widths = [20, 10, 10, 8, 12, 14, 14, 14, 16]
    for c, w in enumerate(sum_widths, 1):
        ws_summary.column_dimensions[get_column_letter(c)].width = w
    ws_summary.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════════════════
    # Cross-version speedup sheet (if multiple solvers)
    # ═══════════════════════════════════════════════════════════════
    if len(solver_entries) >= 2:
        ws_sp = cast(Worksheet, wb.create_sheet("Speedup"))

        sp_headers = ["Solver", "Total Time (s)", "Speedup vs Baseline", "Avg Nodes", "Node Ratio"]
        for c, h in enumerate(sp_headers, 1):
            ws_sp.cell(row=1, column=c, value=h)
        style_header(ws_sp, 1, len(sp_headers))

        baseline_label = solver_labels[0]
        base_time = summary[baseline_label]["total_time"]
        base_nodes = (summary[baseline_label]["total_nodes"] / summary[baseline_label]["count_nodes"]
                      if summary[baseline_label]["count_nodes"] > 0 else 0)

        row = 2
        for s_label in solver_labels:
            s = summary[s_label]
            total_t = s["total_time"]
            avg_n = (s["total_nodes"] / s["count_nodes"]
                     if s["count_nodes"] > 0 else 0)
            sp_val = round(base_time / total_t, 2) if (total_t > 0 and base_time > 0) else None
            node_ratio = round(avg_n / base_nodes, 2) if (base_nodes > 0 and avg_n > 0) else None
            values = [
                s_label,
                round(total_t, 2) if total_t > 0 else None,
                sp_val, round(avg_n, 0) if avg_n > 0 else None,
                node_ratio,
            ]
            for c, v in enumerate(values, 1):
                cell = cast(Cell, ws_sp.cell(row=row, column=c))
                cell.value = v if v is not None else "-"
                style_data_cell(cell)
            row += 1

        sp_widths = [20, 14, 18, 14, 12]
        for c, w in enumerate(sp_widths, 1):
            ws_sp.column_dimensions[get_column_letter(c)].width = w
        ws_sp.freeze_panes = "A2"

    # Remove default sheet if we added real ones
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    # Move Per-Instance Results to first position
    if "Per-Instance Results" in wb.sheetnames:
        wb.move_sheet("Per-Instance Results", offset=0)

    # ── Save ──
    wb.save(xlsx_path)
    print(f"Excel written to: {xlsx_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
