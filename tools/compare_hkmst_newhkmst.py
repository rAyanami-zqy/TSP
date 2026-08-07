#!/usr/bin/env python3
"""Compare Concorde, HKMST, and NEWHKMST on symmetric TSP instances.

The execution model deliberately follows ``tools/compare_strategies.py``:

* instances are processed in parallel, while the three algorithms for one
  instance are run sequentially;
* Concorde uses a fixed seed and a private temporary directory;
* tsp_bb wall time is measured around the complete subprocess invocation;
* successful results and timeouts are cached and the XLSX is rebuilt each run.

Put the compiled executables at:

    solver/HKMST/tsp_bb
    solver/NEWHKMST/tsp_bb

The NEWHKMST executable is invoked with the currently recommended persistent
potential-epoch configuration.  By default the script discovers project
examples plus TSPLIB/National instances with dimension strictly below 200.

Output:

    docs/HKMST-NEWHKMST-Concorde-comparison.xlsx

Usage:

    python3 tools/compare_hkmst_newhkmst.py
    python3 tools/compare_hkmst_newhkmst.py --fresh --repeats 3
    python3 tools/compare_hkmst_newhkmst.py --instances examples/five-city.txt
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import shlex
import statistics
import subprocess
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any

import compare_strategies as reference

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_HKMST = PROJECT_ROOT / "solver" / "HKMST" / "tsp_bb"
DEFAULT_NEWHKMST = PROJECT_ROOT / "solver" / "NEWHKMST" / "tsp_bb"
DEFAULT_CONCORDE = PROJECT_ROOT / "concorde" / "TSP" / "concorde"
DEFAULT_OUTPUT = PROJECT_ROOT / "docs" / "HKMST-NEWHKMST-Concorde-comparison.xlsx"
DEFAULT_CACHE = PROJECT_ROOT / "docs" / "hkmst-newhkmst-comparison-cache.json"
CONCORDE_WORK_DIR = PROJECT_ROOT / "TS"

DEFAULT_TIMEOUT = 1800
DEFAULT_WORKERS = 5
DEFAULT_REPEATS = 1
DEFAULT_MAX_DIMENSION_EXCLUSIVE = 200
DEBUG_INTERVAL = 5_000_000
CONCORDE_SEED = 123
CACHE_SCHEMA = 1

NEWHKMST_DEFAULT_ARGS = (
    "--hk-ascent polyak "
    "--hk-potential-update subtree-adaptive "
    "--hk-update-depth 2 "
    "--hk-update-gap-ratio 0.02 "
    "--hk-update-iterations 16 "
    "--hk-update-budget 5000"
)

# Same source layout as compare_strategies.py, with the requested n < 200 cap.
INSTANCE_SOURCES: tuple[dict[str, Any], ...] = (
    {
        "path": "data/classic/tsplib",
        "recursive": False,
        "skip_dirs": set(),
        # Kept consistent with the reference script's default exclusion.
        "skip_files": {"ulysses22.tsp"},
        "extensions": (".txt", ".tsp"),
    },
    {
        "path": "data/classic/national",
        "recursive": False,
        "skip_dirs": set(),
        "skip_files": set(),
        "extensions": (".txt", ".tsp"),
    },
)

_cache_lock = threading.Lock()
_print_lock = threading.Lock()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare Concorde, HKMST, and NEWHKMST and export XLSX.")
    parser.add_argument("--hkmst", type=Path, default=DEFAULT_HKMST,
                        help="HKMST tsp_bb executable")
    parser.add_argument("--newhkmst", type=Path, default=DEFAULT_NEWHKMST,
                        help="NEWHKMST tsp_bb executable")
    parser.add_argument("--concorde", type=Path, default=DEFAULT_CONCORDE,
                        help="Concorde executable")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT,
                        help="output XLSX path")
    parser.add_argument("--cache", type=Path, default=DEFAULT_CACHE,
                        help="persistent JSON cache path")
    parser.add_argument("--timeout", type=int, default=DEFAULT_TIMEOUT,
                        help="seconds allowed per algorithm and instance")
    parser.add_argument("--workers", type=int, default=DEFAULT_WORKERS,
                        help="number of instances processed concurrently")
    parser.add_argument("--repeats", type=int, default=DEFAULT_REPEATS,
                        help="wall-time runs per algorithm; median is reported")
    parser.add_argument(
        "--max-dimension", type=int,
        default=DEFAULT_MAX_DIMENSION_EXCLUSIVE,
        help="exclusive dimension limit; default 200 means n < 200")
    parser.add_argument("--limit", type=int, default=None,
                        help="run only the first N discovered instances")
    parser.add_argument("--instances", nargs="*", type=Path,
                        help="explicit instance paths instead of discovery")
    parser.add_argument("--hkmst-args", default="",
                        help="additional quoted arguments for HKMST")
    parser.add_argument("--newhkmst-args", default=NEWHKMST_DEFAULT_ARGS,
                        help="arguments for NEWHKMST; replaces the default set")
    parser.add_argument("--fresh", action="store_true",
                        help="ignore and overwrite the existing cache")
    parser.add_argument("--list-instances", action="store_true",
                        help="print selected instances and exit")
    parser.add_argument("--no-xlsx", action="store_true",
                        help="run/cache results without creating XLSX")
    args = parser.parse_args()

    if args.timeout <= 0 or args.workers <= 0 or args.repeats <= 0:
        parser.error("--timeout, --workers, and --repeats must be positive")
    if args.max_dimension <= 3:
        parser.error("--max-dimension must be greater than 3")
    if args.limit is not None and args.limit <= 0:
        parser.error("--limit must be positive")
    return args


def discover_instances(max_dimension: int,
                       explicit: list[Path] | None = None) -> list[tuple[Path, int]]:
    """Return deduplicated instances sorted by dimension then basename."""
    candidates: list[Path] = []
    if explicit is not None:
        candidates = [path.resolve() for path in explicit]
    else:
        for cfg in INSTANCE_SOURCES:
            base = PROJECT_ROOT / cfg["path"]
            if not base.is_dir():
                print(f"Warning: missing instance source: {base}", file=sys.stderr)
                continue
            if cfg["recursive"]:
                for dirpath, dirnames, filenames in os.walk(base):
                    dirnames[:] = sorted(
                        name for name in dirnames
                        if not name.startswith(".") and name not in cfg["skip_dirs"])
                    for filename in sorted(filenames):
                        if (not filename.startswith(".")
                                and filename not in cfg["skip_files"]
                                and filename.endswith(cfg["extensions"])):
                            candidates.append(Path(dirpath) / filename)
            else:
                for path in sorted(base.iterdir()):
                    if (path.is_file() and not path.name.startswith(".")
                            and path.name not in cfg["skip_files"]
                            and path.name.endswith(cfg["extensions"])):
                        candidates.append(path)

    instances: list[tuple[Path, int]] = []
    seen_names: set[str] = set()
    for path in candidates:
        if not path.is_file():
            raise FileNotFoundError(f"instance not found: {path}")
        dimension = reference._parse_dimension(str(path))
        if dimension is None or dimension < 3 or dimension >= max_dimension:
            continue
        # Keep the reference script's basename de-duplication rule.
        if path.name in seen_names:
            continue
        seen_names.add(path.name)
        instances.append((path.resolve(), dimension))
    instances.sort(key=lambda item: (item[1], item[0].name, str(item[0])))
    return instances


def binary_digest(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def experiment_signature(args: argparse.Namespace,
                         hkmst_args: list[str],
                         newhkmst_args: list[str]) -> dict[str, Any]:
    return {
        "schema": CACHE_SCHEMA,
        "timeout": args.timeout,
        "repeats": args.repeats,
        "max_dimension_exclusive": args.max_dimension,
        "concorde_seed": CONCORDE_SEED,
        "hkmst_args": hkmst_args,
        "newhkmst_args": newhkmst_args,
        "binaries": {
            "Concorde": binary_digest(args.concorde),
            "HKMST": binary_digest(args.hkmst),
            "NEWHKMST": binary_digest(args.newhkmst),
        },
    }


def load_cache(path: Path, signature: dict[str, Any], fresh: bool) -> dict[str, Any]:
    empty = {"signature": signature, "results": {}}
    if fresh or not path.is_file():
        return empty
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        print(f"Warning: ignoring unreadable cache: {error}", file=sys.stderr)
        return empty
    if data.get("signature") != signature:
        print("Cache signature changed; starting a fresh experiment cache.",
              file=sys.stderr)
        return empty
    data.setdefault("results", {})
    return data


def save_cache(path: Path, cache: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(cache, indent=2, sort_keys=True, ensure_ascii=False),
        encoding="utf-8")
    temporary.replace(path)


def json_stats(stats: dict[str, Any]) -> dict[str, Any]:
    keep = (
        "cost", "feasible", "elapsed", "elapsed_runs", "timeout", "error",
        "bbnodes", "root_lb", "init_ub", "nodes_created", "nodes_expanded",
        "pruned_bound", "pruned_infeasible",
    )
    return {key: stats.get(key) for key in keep}


def empty_solver_stats() -> dict[str, Any]:
    stats = reference._empty_stats_tspbb()
    stats["elapsed_runs"] = []
    return stats


def run_solver_once(instance: Path, binary: Path, extra_args: list[str],
                    timeout: int, max_dimension: int) -> dict[str, Any]:
    command = [
        str(binary),
        *extra_args,
        "--exact-max-n", str(max_dimension - 1),
        "--debug",
        "--debug-interval", str(DEBUG_INTERVAL),
        str(instance),
    ]
    started = time.perf_counter()
    try:
        completed = subprocess.run(
            command, capture_output=True, text=True, timeout=timeout)
        elapsed = time.perf_counter() - started
        stats = reference.parse_tspbb_stats(completed.stdout)
        stats.update({
            "elapsed": elapsed,
            "timeout": False,
            "error": None,
            "stderr": completed.stderr,
        })
        if completed.returncode != 0:
            message = completed.stderr.strip() or completed.stdout.strip()
            stats["error"] = (
                f"exit {completed.returncode}: {message[-500:]}".strip())
        elif not stats.get("feasible"):
            stats["error"] = "solver produced no parseable optimal result"
        return stats
    except subprocess.TimeoutExpired:
        stats = empty_solver_stats()
        stats["elapsed"] = float(timeout)
        stats["timeout"] = True
        return stats
    except Exception as error:  # noqa: BLE001 - recorded in experiment output
        stats = empty_solver_stats()
        stats["elapsed"] = time.perf_counter() - started
        stats["error"] = str(error)
        return stats


def aggregate_runs(runs: list[dict[str, Any]]) -> dict[str, Any]:
    """Use median wall time and verify repeat outcomes are consistent."""
    result = dict(runs[0])
    elapsed_runs = [float(run.get("elapsed", 0.0)) for run in runs]
    result["elapsed_runs"] = elapsed_runs
    result["elapsed"] = statistics.median(elapsed_runs)

    signatures = {
        (
            bool(run.get("timeout")),
            bool(run.get("error")),
            bool(run.get("feasible")),
            run.get("cost"),
            run.get("nodes_created"),
            run.get("nodes_expanded"),
            run.get("bbnodes"),
        )
        for run in runs
    }
    if len(signatures) != 1:
        result["error"] = "inconsistent result across repeated runs"
        result["feasible"] = False
    return result


def run_repeated(runner, repeats: int) -> dict[str, Any]:
    runs: list[dict[str, Any]] = []
    for _ in range(repeats):
        stats = runner()
        runs.append(stats)
        if stats.get("timeout") or stats.get("error"):
            break
    return aggregate_runs(runs)


def result_status(stats: dict[str, Any], reference_cost: float | None,
                  is_concorde: bool = False) -> str:
    if stats.get("timeout"):
        return "TIMEOUT"
    if stats.get("error"):
        return "ERROR"
    if not stats.get("feasible"):
        return "NO SOLUTION"
    if is_concorde:
        return "OPTIMAL"
    if reference_cost is None:
        return "SOLVED (NO REF)"
    if abs(float(stats["cost"]) - float(reference_cost)) <= 1e-6:
        return "OPTIMAL"
    return "WRONG COST"


def relative_instance_path(path: Path) -> str:
    try:
        return str(path.relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path)


def process_instance(index: int, total: int, instance: Path, dimension: int,
                     args: argparse.Namespace, hkmst_args: list[str],
                     newhkmst_args: list[str], cache: dict[str, Any]) -> dict[str, Any]:
    rel_path = relative_instance_path(instance)
    log = [f"[{index}/{total}] {rel_path} (n={dimension})"]

    def cached_or_run(label: str, runner) -> tuple[dict[str, Any], bool]:
        with _cache_lock:
            cached = cache["results"].get(rel_path, {}).get(label)
        if cached is not None:
            return dict(cached), True
        stats = run_repeated(runner, args.repeats)
        stored = json_stats(stats)
        with _cache_lock:
            cache["results"].setdefault(rel_path, {})[label] = stored
            save_cache(args.cache, cache)
        return stored, False

    concorde, cached = cached_or_run(
        "Concorde", lambda: reference.run_concorde(str(instance), CONCORDE_SEED))
    reference_cost = concorde.get("cost") if concorde.get("feasible") else None
    log.append(
        f"  Concorde: {result_status(concorde, reference_cost, True)} "
        f"cost={concorde.get('cost')} time={concorde.get('elapsed', 0):.6f}s"
        + (" (cached)" if cached else ""))

    hkmst, cached = cached_or_run(
        "HKMST",
        lambda: run_solver_once(
            instance, args.hkmst, hkmst_args, args.timeout, args.max_dimension))
    log.append(
        f"  HKMST: {result_status(hkmst, reference_cost)} "
        f"cost={hkmst.get('cost')} time={hkmst.get('elapsed', 0):.6f}s "
        f"created={hkmst.get('nodes_created')}"
        + (" (cached)" if cached else ""))

    newhkmst, cached = cached_or_run(
        "NEWHKMST",
        lambda: run_solver_once(
            instance, args.newhkmst, newhkmst_args,
            args.timeout, args.max_dimension))
    log.append(
        f"  NEWHKMST: {result_status(newhkmst, reference_cost)} "
        f"cost={newhkmst.get('cost')} time={newhkmst.get('elapsed', 0):.6f}s "
        f"created={newhkmst.get('nodes_created')}"
        + (" (cached)" if cached else ""))

    return {
        "index": index,
        "instance": rel_path,
        "dimension": dimension,
        "Concorde": concorde,
        "HKMST": hkmst,
        "NEWHKMST": newhkmst,
        "log": log,
    }


def write_workbook(results: list[dict[str, Any]], output: Path,
                   args: argparse.Namespace, signature: dict[str, Any],
                   hkmst_args: list[str], newhkmst_args: list[str]) -> None:
    if not HAS_OPENPYXL:
        raise RuntimeError(
            "openpyxl is required for XLSX output; install it with "
            "'python3 -m pip install openpyxl'")

    workbook = Workbook()
    detail = workbook.active
    detail.title = "逐实例对比"
    summary = workbook.create_sheet("汇总", 0)
    config = workbook.create_sheet("运行配置")

    # Palette: neutral Concorde, blue HKMST, teal NEWHKMST.
    navy = "17365D"
    concorde_fill = PatternFill("solid", fgColor="595959")
    hkmst_fill = PatternFill("solid", fgColor="4472C4")
    newhkmst_fill = PatternFill("solid", fgColor="008C95")
    title_fill = PatternFill("solid", fgColor=navy)
    subheader_fill = PatternFill("solid", fgColor="D9EAF7")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    green_font = Font(color="006100")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    red_font = Font(color="9C0006")
    amber_fill = PatternFill("solid", fgColor="FFEB9C")
    amber_font = Font(color="9C6500")
    neutral_fill = PatternFill("solid", fgColor="DDEBF7")
    neutral_font = Font(color="1F4E78")
    white_bold = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E1F2")
    section_border = Border(bottom=Side(style="medium", color="9EADBA"))

    def style_title(sheet, cell_range: str, text: str) -> None:
        sheet.merge_cells(cell_range)
        cell = sheet[cell_range.split(":")[0]]
        cell.value = text
        cell.fill = title_fill
        cell.font = Font(color="FFFFFF", bold=True, size=16)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[cell.row].height = 28

    def apply_percent_rules(sheet, cell_range: str) -> None:
        sheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=green_fill, font=green_font))
        sheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=red_fill, font=red_font))

    # ── Per-instance table ────────────────────────────────────────
    style_title(
        detail, "A1:U1",
        "Concorde / HKMST / NEWHKMST 逐实例精确求解对比")
    group_headers = (
        ("A2:C2", "实例", title_fill),
        ("D2:G2", "Concorde", concorde_fill),
        ("H2:M2", "HKMST", hkmst_fill),
        ("N2:U2", "NEWHKMST", newhkmst_fill),
    )
    for cell_range, text, fill in group_headers:
        detail.merge_cells(cell_range)
        cell = detail[cell_range.split(":")[0]]
        cell.value = text
        cell.fill = fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
    detail.row_dimensions[2].height = 24

    headers = [
        "序号", "实例名", "顶点数",
        "状态", "求解结果", "运行时间(s)", "B&B节点数",
        "状态", "求解结果", "运行时间(s)", "创建节点", "展开节点",
        "时间提升 vs Concorde",
        "状态", "求解结果", "运行时间(s)", "创建节点", "展开节点",
        "时间提升 vs Concorde", "时间提升 vs HKMST", "分支减少 vs HKMST",
    ]
    for column, header in enumerate(headers, 1):
        cell = detail.cell(3, column, header)
        if column <= 3:
            cell.fill = subheader_fill
            cell.font = Font(bold=True, color=navy)
        elif column <= 7:
            cell.fill = concorde_fill
            cell.font = white_bold
        elif column <= 13:
            cell.fill = hkmst_fill
            cell.font = white_bold
        else:
            cell.fill = newhkmst_fill
            cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = section_border
    detail.row_dimensions[3].height = 42

    first_data_row = 4
    for row, result in enumerate(results, first_data_row):
        concorde = result["Concorde"]
        hkmst = result["HKMST"]
        newhkmst = result["NEWHKMST"]
        reference_cost = concorde.get("cost") if concorde.get("feasible") else None
        values = [
            result["index"], result["instance"], result["dimension"],
            result_status(concorde, reference_cost, True),
            concorde.get("cost"), concorde.get("elapsed"), concorde.get("bbnodes"),
            result_status(hkmst, reference_cost),
            hkmst.get("cost"), hkmst.get("elapsed"),
            hkmst.get("nodes_created"), hkmst.get("nodes_expanded"), None,
            result_status(newhkmst, reference_cost),
            newhkmst.get("cost"), newhkmst.get("elapsed"),
            newhkmst.get("nodes_created"), newhkmst.get("nodes_expanded"),
            None, None, None,
        ]
        for column, value in enumerate(values, 1):
            cell = detail.cell(row, column, value)
            cell.alignment = Alignment(
                horizontal="left" if column == 2 else "center",
                vertical="center")
            cell.border = Border(bottom=thin_gray)
        # Positive means faster/fewer and is therefore green.
        detail.cell(
            row, 13,
            f'=IF(AND(D{row}="OPTIMAL",H{row}="OPTIMAL",F{row}>0),'
            f'(F{row}-J{row})/F{row},"")')
        detail.cell(
            row, 19,
            f'=IF(AND(D{row}="OPTIMAL",N{row}="OPTIMAL",F{row}>0),'
            f'(F{row}-P{row})/F{row},"")')
        detail.cell(
            row, 20,
            f'=IF(AND(H{row}="OPTIMAL",N{row}="OPTIMAL",J{row}>0),'
            f'(J{row}-P{row})/J{row},"")')
        detail.cell(
            row, 21,
            f'=IF(AND(H{row}="OPTIMAL",N{row}="OPTIMAL",K{row}>0),'
            f'(K{row}-Q{row})/K{row},"")')

    last_data_row = max(first_data_row, first_data_row + len(results) - 1)
    detail.auto_filter.ref = f"A3:U{last_data_row}"
    detail.freeze_panes = "D4"
    detail.sheet_view.showGridLines = False
    detail.sheet_properties.pageSetUpPr.fitToPage = True
    detail.page_setup.orientation = "landscape"
    detail.page_setup.fitToWidth = 1
    detail.page_setup.fitToHeight = 0
    detail.print_title_rows = "1:3"

    for column in (5, 9, 15):
        detail.cell(3, column).number_format = "0.########"
        for row in range(first_data_row, last_data_row + 1):
            detail.cell(row, column).number_format = "0.########"
    for column in (6, 10, 16):
        for row in range(first_data_row, last_data_row + 1):
            detail.cell(row, column).number_format = "0.000000"
    for column in (7, 11, 12, 17, 18):
        for row in range(first_data_row, last_data_row + 1):
            detail.cell(row, column).number_format = "#,##0"
    for column in (13, 19, 20, 21):
        for row in range(first_data_row, last_data_row + 1):
            detail.cell(row, column).number_format = "0.00%"
        apply_percent_rules(
            detail, f"{get_column_letter(column)}{first_data_row}:"
                    f"{get_column_letter(column)}{last_data_row}")

    for status_column in (4, 8, 14):
        status_range = (
            f"{get_column_letter(status_column)}{first_data_row}:"
            f"{get_column_letter(status_column)}{last_data_row}")
        detail.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'{get_column_letter(status_column)}{first_data_row}="OPTIMAL"'],
                fill=green_fill, font=green_font))
        detail.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'OR({get_column_letter(status_column)}{first_data_row}="ERROR",'
                         f'{get_column_letter(status_column)}{first_data_row}="WRONG COST")'],
                fill=red_fill, font=red_font))
        detail.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'{get_column_letter(status_column)}{first_data_row}="TIMEOUT"'],
                fill=amber_fill, font=amber_font))
        detail.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'{get_column_letter(status_column)}{first_data_row}="SOLVED (NO REF)"'],
                fill=neutral_fill, font=neutral_font))

    widths = {
        "A": 7, "B": 42, "C": 9,
        "D": 13, "E": 14, "F": 15, "G": 15,
        "H": 13, "I": 14, "J": 15, "K": 14, "L": 14, "M": 19,
        "N": 13, "O": 14, "P": 15, "Q": 14, "R": 14,
        "S": 19, "T": 18, "U": 18,
    }
    for column, width in widths.items():
        detail.column_dimensions[column].width = width

    # ── Summary table ─────────────────────────────────────────────
    style_title(summary, "A1:N1", "HKMST 与 NEWHKMST 实验汇总")
    summary["A2"] = (
        "提升率=(基准-当前)/基准；正数表示更快或分支更少（绿色），"
        "负数表示退化（红色）。时间为完整子进程 wall time；"
        "分支数采用 tsp_bb 的 Nodes created。汇总提升率只统计双方均返回"
        "Concorde最优值的可比实例。")
    summary.merge_cells("A2:N2")
    summary["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    summary["A2"].font = Font(color="595959", italic=True)
    summary.row_dimensions[2].height = 34
    summary_headers = [
        "算法", "完成求解", "与Concorde一致", "超时", "错误", "错误结果",
        "总时间(s)", "平均时间(s)", "中位时间(s)", "总节点/分支",
        "总展开节点", "时间提升 vs Concorde", "时间提升 vs HKMST",
        "分支减少 vs HKMST",
    ]
    for column, header in enumerate(summary_headers, 1):
        cell = summary.cell(3, column, header)
        cell.fill = title_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    summary.row_dimensions[3].height = 42

    detail_refs = {
        "Concorde": {"status": "D", "time": "F", "nodes": "G", "expanded": None},
        "HKMST": {"status": "H", "time": "J", "nodes": "K", "expanded": "L"},
        "NEWHKMST": {"status": "N", "time": "P", "nodes": "Q", "expanded": "R"},
    }

    def paired_improvement_formula(
            base_value_column: str, current_value_column: str,
            base_status_column: str, current_status_column: str) -> str:
        base_values = (
            f"'逐实例对比'!${base_value_column}$4:"
            f"${base_value_column}${last_data_row}")
        current_values = (
            f"'逐实例对比'!${current_value_column}$4:"
            f"${current_value_column}${last_data_row}")
        base_status = (
            f"'逐实例对比'!${base_status_column}$4:"
            f"${base_status_column}${last_data_row}")
        current_status = (
            f"'逐实例对比'!${current_status_column}$4:"
            f"${current_status_column}${last_data_row}")
        criteria = (
            f'{base_status},"OPTIMAL",{current_status},"OPTIMAL"')
        base_sum = f"SUMIFS({base_values},{criteria})"
        current_sum = f"SUMIFS({current_values},{criteria})"
        return f'=IFERROR(({base_sum}-{current_sum})/{base_sum},"")'

    for summary_row, algorithm in enumerate(("Concorde", "HKMST", "NEWHKMST"), 4):
        refs = detail_refs[algorithm]
        status_range = f"'逐实例对比'!${refs['status']}$4:${refs['status']}${last_data_row}"
        time_range = f"'逐实例对比'!${refs['time']}$4:${refs['time']}${last_data_row}"
        node_range = f"'逐实例对比'!${refs['nodes']}$4:${refs['nodes']}${last_data_row}"
        summary.cell(summary_row, 1, algorithm)
        summary.cell(summary_row, 2,
                     f'=COUNTIF({status_range},"OPTIMAL")+'
                     f'COUNTIF({status_range},"WRONG COST")+'
                     f'COUNTIF({status_range},"SOLVED (NO REF)")')
        summary.cell(summary_row, 3, f'=COUNTIF({status_range},"OPTIMAL")')
        summary.cell(summary_row, 4, f'=COUNTIF({status_range},"TIMEOUT")')
        summary.cell(summary_row, 5, f'=COUNTIF({status_range},"ERROR")')
        summary.cell(summary_row, 6, f'=COUNTIF({status_range},"WRONG COST")')
        summary.cell(summary_row, 7, f'=SUM({time_range})')
        summary.cell(summary_row, 8, f'=IFERROR(AVERAGE({time_range}),"")')
        summary.cell(summary_row, 9, f'=IFERROR(MEDIAN({time_range}),"")')
        summary.cell(summary_row, 10, f'=SUM({node_range})')
        if refs["expanded"] is not None:
            expanded_range = (
                f"'逐实例对比'!${refs['expanded']}$4:"
                f"${refs['expanded']}${last_data_row}")
            summary.cell(summary_row, 11, f'=SUM({expanded_range})')
        if algorithm == "HKMST":
            summary.cell(
                summary_row, 12,
                paired_improvement_formula("F", "J", "D", "H"))
        elif algorithm == "NEWHKMST":
            summary.cell(
                summary_row, 12,
                paired_improvement_formula("F", "P", "D", "N"))
            summary.cell(
                summary_row, 13,
                paired_improvement_formula("J", "P", "H", "N"))
            summary.cell(
                summary_row, 14,
                paired_improvement_formula("K", "Q", "H", "N"))
        for column in range(1, 15):
            cell = summary.cell(summary_row, column)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin_gray)

    for column in (7, 8, 9):
        for row in range(4, 7):
            summary.cell(row, column).number_format = "0.000000"
    for column in (10, 11):
        for row in range(4, 7):
            summary.cell(row, column).number_format = "#,##0"
    for column in (12, 13, 14):
        for row in range(4, 7):
            summary.cell(row, column).number_format = "0.00%"
        apply_percent_rules(
            summary, f"{get_column_letter(column)}4:{get_column_letter(column)}6")
    summary.freeze_panes = "B4"
    summary.sheet_view.showGridLines = False
    for column in range(1, 15):
        summary.column_dimensions[get_column_letter(column)].width = (
            20 if column == 1 else 16)

    # ── Audit/configuration sheet ─────────────────────────────────
    style_title(config, "A1:C1", "实验运行配置")
    config.append(["项目", "值", "说明"])
    config_rows = [
        ("生成时间", datetime.now().astimezone().isoformat(timespec="seconds"), "本地时区"),
        ("实例数量", len(results), "每个实例在逐实例表中占一行"),
        ("顶点限制", f"n < {args.max_dimension}", "默认仅测试小于200顶点"),
        ("单次超时", args.timeout, "秒/算法/实例"),
        ("并行实例数", args.workers, "同一实例的三种算法仍串行"),
        ("单实例运行顺序", "Concorde -> HKMST -> NEWHKMST",
         "与 compare_strategies.py 的执行方式一致"),
        ("重复次数", args.repeats, "运行时间取中位数"),
        ("Concorde seed", CONCORDE_SEED, "固定随机种子"),
        ("Concorde", str(args.concorde), signature["binaries"]["Concorde"]),
        ("HKMST", str(args.hkmst), signature["binaries"]["HKMST"]),
        ("HKMST参数", " ".join(hkmst_args) or "(无额外参数)", ""),
        ("NEWHKMST", str(args.newhkmst), signature["binaries"]["NEWHKMST"]),
        ("NEWHKMST参数", " ".join(newhkmst_args), "持久子树势更新策略"),
        ("缓存文件", str(args.cache), "二进制或参数变化时自动失效"),
    ]
    for row in config_rows:
        config.append(list(row))
    for cell in config[2]:
        cell.fill = title_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center")
    for row in config.iter_rows(min_row=3, max_row=config.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin_gray)
    config.column_dimensions["A"].width = 22
    config.column_dimensions["B"].width = 90
    config.column_dimensions["C"].width = 72
    config.freeze_panes = "A3"
    config.sheet_view.showGridLines = False

    # Recalculate formulas when opened in Excel/LibreOffice.
    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True
        calculation.calcMode = "auto"

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def validate_executable(path: Path, label: str) -> None:
    if not path.is_file():
        raise FileNotFoundError(f"{label} executable not found: {path}")
    if not os.access(path, os.X_OK):
        raise PermissionError(f"{label} is not executable: {path}")


def main() -> int:
    args = parse_args()
    args.hkmst = args.hkmst.resolve()
    args.newhkmst = args.newhkmst.resolve()
    args.concorde = args.concorde.resolve()
    args.output = args.output.resolve()
    args.cache = args.cache.resolve()

    instances = discover_instances(args.max_dimension, args.instances)
    if args.limit is not None:
        instances = instances[:args.limit]
    if args.list_instances:
        for path, dimension in instances:
            print(f"{dimension:3d}  {relative_instance_path(path)}")
        print(f"Total: {len(instances)}")
        return 0
    if not instances:
        print("Error: no instances selected", file=sys.stderr)
        return 2

    try:
        validate_executable(args.concorde, "Concorde")
        validate_executable(args.hkmst, "HKMST")
        validate_executable(args.newhkmst, "NEWHKMST")
    except (FileNotFoundError, PermissionError) as error:
        print(f"Error: {error}", file=sys.stderr)
        print(
            "Copy compiled tsp_bb files to solver/HKMST/tsp_bb and "
            "solver/NEWHKMST/tsp_bb.", file=sys.stderr)
        return 2
    if not args.no_xlsx and not HAS_OPENPYXL:
        print(
            "Error: openpyxl is required for the formatted XLSX output. "
            "Install it with: python3 -m pip install openpyxl",
            file=sys.stderr)
        return 2

    hkmst_args = shlex.split(args.hkmst_args)
    newhkmst_args = shlex.split(args.newhkmst_args)
    reference.CONCORDE = str(args.concorde)
    reference.CONCORDE_OUT_DIR = str(CONCORDE_WORK_DIR)
    reference.TIMEOUT = args.timeout
    reference.CONCORDE_SEED = CONCORDE_SEED

    signature = experiment_signature(args, hkmst_args, newhkmst_args)
    cache = load_cache(args.cache, signature, args.fresh)
    print(
        f"Selected {len(instances)} instances, workers={args.workers}, "
        f"repeats={args.repeats}, timeout={args.timeout}s",
        file=sys.stderr)
    print(f"HKMST: {args.hkmst}", file=sys.stderr)
    print(f"NEWHKMST: {args.newhkmst}", file=sys.stderr)
    print(f"NEWHKMST args: {' '.join(newhkmst_args)}", file=sys.stderr)

    indexed_results: dict[int, dict[str, Any]] = {}
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(
                process_instance, index, len(instances), path, dimension,
                args, hkmst_args, newhkmst_args, cache): index
            for index, (path, dimension) in enumerate(instances, 1)
        }
        for future in as_completed(futures):
            result = future.result()
            indexed_results[result["index"]] = result
            with _print_lock:
                for line in result["log"]:
                    print(line, file=sys.stderr)

    results = [indexed_results[index] for index in sorted(indexed_results)]
    if not args.no_xlsx:
        write_workbook(
            results, args.output, args, signature, hkmst_args, newhkmst_args)
        print(f"Excel written to: {args.output}", file=sys.stderr)
    else:
        print("XLSX export skipped (--no-xlsx).", file=sys.stderr)
    print(f"Cache written to: {args.cache}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
